import os
import re
import hashlib
import hmac
import base64
import json
import io
from datetime import date, timedelta, datetime
from typing import Optional

import requests
from fastapi import FastAPI, Request, HTTPException, Depends, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from pydantic import BaseModel
from dotenv import load_dotenv
import openpyxl
import anthropic

load_dotenv()

LINE_TOKEN = os.getenv("LINE_CHANNEL_ACCESS_TOKEN", "")
LINE_SECRET = os.getenv("LINE_CHANNEL_SECRET", "")
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
LIFF_ID = os.getenv("LIFF_ID", "")
API_BASE = os.getenv("API_BASE", "")
PORT = int(os.getenv("PORT", "10001"))
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")

_anthropic_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else None

SUPABASE_HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

SHIFT_TYPES = {"7-3", "9-5", "10-6", "12-8", "3-11", "11-7", "其他", "休"}

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DOCS_DIR = os.path.join(os.path.dirname(__file__), "docs")

# ── LINE helpers ──────────────────────────────────────────────────────────────

def reply_message(reply_token: str, text: str):
    requests.post(
        "https://api.line.me/v2/bot/message/reply",
        headers={"Authorization": f"Bearer {LINE_TOKEN}", "Content-Type": "application/json"},
        json={"replyToken": reply_token, "messages": [{"type": "text", "text": text}]},
        timeout=5,
    )


def push_message(line_user_id: str, text: str):
    requests.post(
        "https://api.line.me/v2/bot/message/push",
        headers={"Authorization": f"Bearer {LINE_TOKEN}", "Content-Type": "application/json"},
        json={"to": line_user_id, "messages": [{"type": "text", "text": text}]},
        timeout=5,
    )


# ── Supabase: user helpers ─────────────────────────────────────────────────────

def _sb(path: str, method="GET", params=None, json_body=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    r = requests.request(method, url, headers=SUPABASE_HEADERS, params=params, json=json_body, timeout=10)
    if r.status_code >= 400:
        return None
    return r.json()


def _sb_all(path: str, params: list, page_size: int = 1000) -> list:
    """分頁撈取，繞過 Supabase max-rows 限制。"""
    all_rows = []
    offset = 0
    base_params = [(k, v) for k, v in params if k not in ("limit", "offset")]
    while True:
        paged = base_params + [("limit", str(page_size)), ("offset", str(offset))]
        rows = _sb(path, params=paged)
        if not rows:
            break
        all_rows.extend(rows)
        if len(rows) < page_size:
            break
        offset += page_size
    return all_rows


def get_user_by_line_id(line_user_id: str):
    rows = _sb("nurses", params={"line_user_id": f"eq.{line_user_id}", "limit": "1"})
    return rows[0] if rows else None


def get_user_by_id(user_id: str):
    rows = _sb("nurses", params={"id": f"eq.{user_id}", "limit": "1"})
    return rows[0] if rows else None


def get_user_by_name(name: str):
    rows = _sb("nurses", params={"name": f"eq.{name}", "limit": "1"})
    return rows[0] if rows else None


def get_all_users():
    return _sb("nurses", params={"order": "sort_order.asc.nullslast,name.asc"}) or []


def get_reviewers():
    return _sb("nurses", params={"role": "in.(manager,admin)"}) or []


# ── Supabase: schedule helpers ─────────────────────────────────────────────────

def get_schedules_by_user_date_range(user_id: str, start: str, end: str):
    params = [
        ("user_id", f"eq.{user_id}"),
        ("schedule_date", f"gte.{start}"),
        ("schedule_date", f"lte.{end}"),
        ("status", "eq.active"),
        ("order", "schedule_date"),
        ("limit", "10000"),
    ]
    return _sb("schedules", params=params) or []


def _get_schedules_range(start: str, end: str, user_id: str = None):
    params = [
        ("schedule_date", f"gte.{start}"),
        ("schedule_date", f"lte.{end}"),
        ("status", "eq.active"),
        ("order", "schedule_date,user_id"),
        ("select", "id,user_id,schedule_date,shift_type,area,notes,source_version"),
    ]
    if user_id:
        params.append(("user_id", f"eq.{user_id}"))
    return _sb_all("schedules", params)


def get_schedules_by_date(schedule_date: str):
    return _sb("schedules", params={
        "schedule_date": f"eq.{schedule_date}",
        "status": "eq.active",
        "order": "shift_type,user_id",
        "select": "id,user_id,schedule_date,shift_type,area,notes",
        "limit": "10000",
    }) or []


def get_schedule_by_id(schedule_id: str):
    rows = _sb("schedules", params={"id": f"eq.{schedule_id}", "limit": "1"})
    return rows[0] if rows else None


def upsert_schedules(records: list):
    return requests.post(
        f"{SUPABASE_URL}/rest/v1/schedules",
        headers={**SUPABASE_HEADERS, "Prefer": "return=minimal"},
        json=records,
        timeout=30,
    ).status_code < 400


def delete_schedules_by_version(version: str) -> bool:
    if not version:
        return True
    return requests.delete(
        f"{SUPABASE_URL}/rest/v1/schedules",
        headers=SUPABASE_HEADERS,
        params={"source_version": f"eq.{version}"},
        timeout=30,
    ).status_code < 400


def update_schedule_status(schedule_id: str, status: str):
    requests.patch(
        f"{SUPABASE_URL}/rest/v1/schedules",
        headers=SUPABASE_HEADERS,
        params={"id": f"eq.{schedule_id}"},
        json={"status": status},
        timeout=10,
    )


# ── Supabase: OT priority helpers ──────────────────────────────────────────────

def get_ot_priority_by_date(priority_date: str, shift_type: str = None):
    params = {
        "priority_date": f"eq.{priority_date}",
        "status": "eq.active",
        "order": "shift_type,priority_order",
        "select": "id,priority_date,user_id,shift_type,priority_order,source_version",
        "limit": "10000",
    }
    if shift_type:
        params["shift_type"] = f"eq.{shift_type}"
    return _sb("ot_priority", params=params) or []


def get_ot_priority_by_id(ot_priority_id: str):
    rows = _sb("ot_priority", params={"id": f"eq.{ot_priority_id}", "limit": "1"})
    return rows[0] if rows else None


def get_ot_priority_range(start: str, end: str, user_id: str = None):
    params = [
        ("priority_date", f"gte.{start}"),
        ("priority_date", f"lte.{end}"),
        ("status", "eq.active"),
        ("order", "priority_date,shift_type,priority_order"),
        ("select", "id,priority_date,user_id,shift_type,priority_order,source_version"),
        ("limit", "10000"),
    ]
    if user_id:
        params.append(("user_id", f"eq.{user_id}"))
    return _sb("ot_priority", params=params) or []


def upsert_ot_priority(records: list):
    return requests.post(
        f"{SUPABASE_URL}/rest/v1/ot_priority",
        headers={**SUPABASE_HEADERS, "Prefer": "return=minimal"},
        json=records,
        timeout=30,
    ).status_code < 400


def delete_ot_priority_by_version(version: str) -> bool:
    if not version:
        return True
    return requests.delete(
        f"{SUPABASE_URL}/rest/v1/ot_priority",
        headers=SUPABASE_HEADERS,
        params={"source_version": f"eq.{version}"},
        timeout=30,
    ).status_code < 400


def update_ot_priority_status(ot_priority_id: str, status: str):
    requests.patch(
        f"{SUPABASE_URL}/rest/v1/ot_priority",
        headers=SUPABASE_HEADERS,
        params={"id": f"eq.{ot_priority_id}"},
        json={"status": status},
        timeout=10,
    )


# ── Supabase: swap request helpers ────────────────────────────────────────────

def create_swap_request(data: dict):
    r = requests.post(
        f"{SUPABASE_URL}/rest/v1/swap_requests",
        headers=SUPABASE_HEADERS,
        json=data,
        timeout=10,
    )
    if r.status_code < 400:
        result = r.json()
        return result[0] if result else None
    return None


def get_swap_request_by_id(request_id: str):
    rows = _sb("swap_requests", params={"id": f"eq.{request_id}", "limit": "1"})
    return rows[0] if rows else None


def get_swap_requests_by_user(user_id: str):
    return _sb("swap_requests", params={
        "or": f"(requester_id.eq.{user_id},target_user_id.eq.{user_id})",
        "order": "created_at.desc",
        "limit": "20",
    }) or []


def get_pending_swap_requests():
    return _sb("swap_requests", params={
        "request_type": "eq.shift",
        "status": "eq.pending_admin",
        "order": "created_at",
    }) or []


def update_swap_request(request_id: str, fields: dict):
    fields["updated_at"] = datetime.utcnow().isoformat()
    requests.patch(
        f"{SUPABASE_URL}/rest/v1/swap_requests",
        headers=SUPABASE_HEADERS,
        params={"id": f"eq.{request_id}"},
        json=fields,
        timeout=10,
    )


def _get_conflicting_swap_requests(schedule_ids: list, ot_priority_ids: list):
    """Find pending requests that conflict with the given slot IDs."""
    results = []
    active_statuses = ["submitted", "pending_peer"]

    for sid in schedule_ids:
        for field in ["schedule_id", "target_schedule_id"]:
            rows = _sb("swap_requests", params={
                field: f"eq.{sid}",
                "status": f"in.({','.join(active_statuses)})",
            }) or []
            results.extend(rows)

    for oid in ot_priority_ids:
        for field in ["ot_priority_id", "target_ot_priority_id"]:
            rows = _sb("swap_requests", params={
                field: f"eq.{oid}",
                "status": f"in.({','.join(active_statuses)})",
            }) or []
            results.extend(rows)

    seen = set()
    unique = []
    for r in results:
        if r["id"] not in seen:
            seen.add(r["id"])
            unique.append(r)
    return unique


def auto_reject_conflicts(exclude_request_id: str, schedule_ids: list, ot_priority_ids: list):
    """Auto-reject conflicting pending requests and notify affected users."""
    conflicts = _get_conflicting_swap_requests(schedule_ids, ot_priority_ids)
    for req in conflicts:
        if req["id"] == exclude_request_id:
            continue
        update_swap_request(req["id"], {"status": "conflict_rejected"})
        msg = "您的換班申請因另一筆申請已進入審核流程，已自動取消。"
        if req.get("requester_id"):
            requester = get_user_by_id(req["requester_id"])
            if requester and requester.get("line_user_id"):
                push_message(requester["line_user_id"], msg)
        if req.get("target_user_id"):
            target = get_user_by_id(req["target_user_id"])
            if target and target.get("line_user_id"):
                push_message(target["line_user_id"], msg)


# ── Excel parsing ──────────────────────────────────────────────────────────────

def infer_year_month(version: str):
    """從檔名版本字串萃取年月，格式如 '20260414'；失敗回傳今天。"""
    m = re.search(r"(20\d{2})(\d{2})\d{2}", version)
    if m:
        return int(m.group(1)), int(m.group(2))
    today = datetime.today()
    return today.year, today.month


def parse_date_header(raw: str, year: int, month: int) -> str | None:
    """將 '5/1\n五' 解析為 '2026-05-01'。"""
    part = str(raw).split("\n")[0].strip()
    m = re.match(r"(\d{1,2})[/\-](\d{1,2})", part)
    if not m:
        return None
    return f"{year}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"


_SHIFT_NORMALIZE = {
    '7~3': '7-3', '10~6': '10-6', '3~11': '3-11',
    '12~8': '12-8', '9~5': '9-5', '11~7': '11-7', '11~7a': '11-7',
}
_OFF_SHIFTS_KNOWN = {'公休', '休假', '例假', '國定假日', '休', '例休', '補休', 'off', '國定', '休息日', '排休'}
_SHIFT_START_MAP = {7: '7-3', 9: '9-5', 10: '10-6', 12: '12-8', 3: '3-11', 11: '11-7'}


def _normalize_shift_code(raw: str) -> str:
    base = re.split(r'[np]', raw.strip().strip('.'))[0].strip()
    return _SHIFT_NORMALIZE.get(base, base)


def detect_known_format(wb) -> str:
    """偵測是否為三種已知醫院班表格式。回傳 'ot_before' | 'ot_after' | 'pure'。"""
    if '7_3P' in wb.sheetnames or '10_6班' in wb.sheetnames or '值班列印全' in wb.sheetnames:
        return 'ot_before'
    if '本月' in wb.sheetnames:
        ws = wb['本月']
        for row in ws.iter_rows(min_row=6, max_row=10, values_only=True):
            for v in (row[2:10] if len(row) > 2 else []):
                if isinstance(v, str) and '--' in v:
                    return 'ot_after'
    return 'pure'


def _find_date_cols(ws) -> dict:
    """掃描 Row 4（1-indexed）找出所有 datetime 欄，回傳 {col_index: date_str}。"""
    date_cols = {}
    for row in ws.iter_rows(min_row=4, max_row=4, values_only=True):
        for ci, v in enumerate(row):
            if isinstance(v, (date, datetime)):
                year = v.year if isinstance(v, datetime) else v.year
                if year < 2000:  # 過濾 Excel 公式產生的假日期（如 1900-07-02）
                    continue
                ds = v.strftime('%Y-%m-%d') if isinstance(v, datetime) else v.isoformat()
                date_cols[ci] = ds
    return date_cols


def _parse_vnhc_wide(ws, version: str, strip_ot_code: bool = False) -> list:
    """解析純班別 / 加班前 的「本月」sheet（wide 格式）。"""
    date_cols = _find_date_cols(ws)
    if not date_cols:
        return []
    schedules = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not any(row):
            continue
        name = str(row[1] or '').strip() if len(row) > 1 else ''
        if not name or name in ('日期', '姓名'):
            continue
        for ci, date_str in date_cols.items():
            if ci >= len(row):
                continue
            raw = str(row[ci] or '').strip()
            if not raw:
                continue
            shift = _normalize_shift_code(raw) if strip_ot_code else _SHIFT_NORMALIZE.get(raw, raw)
            if not shift or shift.lower() in _OFF_SHIFTS_KNOWN or shift in OFF_SHIFTS:
                continue
            if shift not in SHIFT_TYPES:
                shift = '其他'
            schedules.append({
                'date_str': date_str,
                'name': name,
                'shift_type': shift,
                'area': '',
                'source_version': version,
            })
    return schedules


def _parse_vnhc_ot_priorities(ws, version: str) -> list:
    """從本月 sheet 的班別代碼（如 7~3n35、3~11p8）提取所有班別加班順位，涵蓋全月所有班別。"""
    date_cols = _find_date_cols(ws)
    results = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not any(row):
            continue
        name = str(row[1] or '').strip() if len(row) > 1 else ''
        if not name or name in ('日期', '姓名'):
            continue
        for ci, date_str in date_cols.items():
            if ci >= len(row):
                continue
            raw = str(row[ci] or '').strip().strip('.')
            if not raw:
                continue
            m = re.match(r'^(.+?)[np](\d+)$', raw)
            if not m:
                continue
            base, priority = m.group(1), int(m.group(2))
            shift_type = _SHIFT_NORMALIZE.get(base, base)
            results.append({
                'date_str': date_str,
                'priority_order': priority,
                'name': name,
                'shift_type': shift_type,
                'source_version': version,
            })
    return results


def _parse_7_3P_sheet(ws, version: str) -> list:
    """解析加班前「7_3P」sheet → 每日 7-3 加班順位清單。"""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if not rows:
        return []
    # Row 1: 日期欄在奇數 col index (1, 3, 5...)
    date_row = rows[0]
    col_dates = {}  # col_index → date_str
    for ci, v in enumerate(date_row):
        if isinstance(v, (date, datetime)):
            ds = v.strftime('%Y-%m-%d') if isinstance(v, datetime) else v.isoformat()
            col_dates[ci] = ds

    results = []
    for row in rows[2:]:  # skip Row 2 (星期)
        if not any(row):
            continue
        # 奇數 col 為順位碼, 偶數+1 col 為護理師名字; 結構: n01|name1|n01|name2...
        # 實際: col0=碼, col1=名, col2=碼, col3=名, ...
        # 每兩欄為一個日期的資料
        for date_col_idx, date_str in col_dates.items():
            code_col = date_col_idx - 1  # 順位碼在日期欄前一欄
            name_col = date_col_idx
            if code_col < 0 or code_col >= len(row):
                continue
            code = str(row[code_col] or '').strip()
            name = str(row[name_col] or '').strip() if name_col < len(row) else ''
            if not code or not name:
                continue
            # 解析順位碼: n01→1, n35→35, S1→101, S2→102, Y1→201, Y2→202
            m = re.match(r'n(\d+)$', code, re.IGNORECASE)
            if m:
                order = int(m.group(1))
            elif re.match(r'S(\d+)$', code, re.IGNORECASE):
                order = 100 + int(re.match(r'S(\d+)$', code, re.IGNORECASE).group(1))
            elif re.match(r'Y(\d+)$', code, re.IGNORECASE):
                order = 200 + int(re.match(r'Y(\d+)$', code, re.IGNORECASE).group(1))
            else:
                continue
            results.append({
                'date_str': date_str,
                'priority_order': order,
                'name': name,
                'shift_type': '7-3',
                'source_version': version,
            })
    return results


def _parse_duty_print_full_sheet(ws, version: str, exclude_shifts: set = None) -> list:
    """解析「值班列印全」sheet → 各班別每日加班順位清單。
    col 0: {班別}P{順位}（如 12~8P1、3~11P3、11~7aP1）
    col 1-N: 護理師名字（含字母前後綴，如 C林艾昀S → 林艾昀）
    """
    _VALID_SHIFTS = set(_SHIFT_NORMALIZE.values())
    label_pat = re.compile(r'^(.+?)P(\d+)$')
    chinese_pat = re.compile(r'[一-鿿㐀-䶿]+')

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if not rows:
        return []

    # Row 1: 找連續日期欄（col 1 開始，遇到 None 就停）
    col_dates = {}
    for ci, v in enumerate(rows[0][1:], start=1):
        if isinstance(v, (date, datetime)):
            col_dates[ci] = v.strftime('%Y-%m-%d') if isinstance(v, datetime) else v.isoformat()
        elif col_dates:
            break  # 日期欄結束

    results = []
    for row in rows[2:]:  # skip Row 1（日期）和 Row 2（星期）
        if not row or not row[0]:
            continue
        m = label_pat.match(str(row[0]).strip())
        if not m:
            continue
        shift_code = m.group(1)
        priority_order = int(m.group(2))
        shift_type = _SHIFT_NORMALIZE.get(shift_code, shift_code.replace('~', '-'))
        if shift_type not in _VALID_SHIFTS:
            continue
        if exclude_shifts and shift_type in exclude_shifts:
            continue
        for ci, date_str in col_dates.items():
            if ci >= len(row):
                continue
            raw = str(row[ci] or '').strip()
            if not raw:
                continue
            # 取漢字部分，去除字母前後綴（如 C林艾昀S → 林艾昀）
            found = chinese_pat.findall(raw)
            name = ''.join(found)
            if not name:
                continue
            results.append({
                'date_str': date_str,
                'priority_order': priority_order,
                'name': name,
                'shift_type': shift_type,
                'source_version': version,
            })
    return results


def _parse_10_6_sheet(ws, version: str) -> list:
    """解析加班前「10_6班」sheet → 每日 10-6 加班順位清單。"""
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if not rows:
        return []
    # Row 1: col 1+ 為日期
    date_row = rows[0]
    col_dates = {}  # col_index → date_str
    for ci, v in enumerate(date_row):
        if ci == 0:
            continue
        if isinstance(v, (date, datetime)):
            ds = v.strftime('%Y-%m-%d') if isinstance(v, datetime) else v.isoformat()
            col_dates[ci] = ds

    results = []
    for row in rows[2:]:  # skip Row 2 (兩頭班/星期)
        if not any(row):
            continue
        code = str(row[0] or '').strip()
        mp = re.match(r'p(\d+)$', code, re.IGNORECASE)
        if not mp:
            continue
        order = int(mp.group(1))
        for ci, date_str in col_dates.items():
            if ci >= len(row):
                continue
            name = str(row[ci] or '').strip()
            if not name:
                continue
            results.append({
                'date_str': date_str,
                'priority_order': order,
                'name': name,
                'shift_type': '10-6',
                'source_version': version,
            })
    return results


def _parse_check_sheet(ws) -> dict:
    """解析加班後「check」sheet，回傳 {(name, date_str): actual_hours}。"""
    date_cols = _find_date_cols(ws)
    result = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not any(row):
            continue
        name = str(row[1] or '').strip() if len(row) > 1 else ''
        if not name or name in ('日期', '姓名'):
            continue
        for ci, date_str in date_cols.items():
            if ci >= len(row):
                continue
            v = row[ci]
            if v is None or v == 0:
                continue
            try:
                hours = float(v)
                if hours > 0:
                    result[(name, date_str)] = hours
            except (TypeError, ValueError):
                pass
    return result


def _parse_shift_start(raw: str) -> int | None:
    """從 '7--6', '7--2休1' 等格式取出起始小時。"""
    m = re.match(r'(\d+)--', raw)
    return int(m.group(1)) if m else None


def _parse_ot_after(wb, version: str) -> dict:
    """解析加班後格式，只回傳 {schedules}。
    overtime_records 由加班申報系統管理，不從 Excel 匯入。"""
    if '本月' not in wb.sheetnames:
        return {'schedules': []}

    date_cols = _find_date_cols(wb['本月'])
    ws = wb['本月']
    schedules = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        if not any(row):
            continue
        name = str(row[1] or '').strip() if len(row) > 1 else ''
        if not name or name in ('日期', '姓名'):
            continue
        for ci, date_str in date_cols.items():
            if ci >= len(row):
                continue
            raw = str(row[ci] or '').strip()
            if not raw or raw.lower() in _OFF_SHIFTS_KNOWN:
                continue
            start_h = _parse_shift_start(raw)
            shift_type = _SHIFT_START_MAP.get(start_h, '其他') if start_h is not None else '其他'
            schedules.append({
                'date_str': date_str,
                'name': name,
                'shift_type': shift_type,
                'area': '',
                'source_version': version,
            })
    return {'schedules': schedules}


def upsert_overtime_records(records: list) -> bool:
    return requests.post(
        f"{SUPABASE_URL}/rest/v1/overtime_records",
        headers={**SUPABASE_HEADERS, "Prefer": "resolution=ignore-duplicates,return=minimal"},
        json=records,
        timeout=30,
    ).status_code < 400


def detect_excel_format(wb, year: int, month: int) -> dict | None:
    """用 Claude 判讀 Excel 結構，失敗回傳 None。"""
    print(f"[detect_excel_format] client={'ok' if _anthropic_client else 'None'}, key_set={bool(ANTHROPIC_API_KEY)}")
    if not _anthropic_client:
        return None

    def _rows_preview(ws, n=4):
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=n, values_only=True)):
            rows.append([str(v) if v is not None else None for v in row])
        return rows

    per_sheet = {name: _rows_preview(wb[name]) for name in wb.sheetnames}
    prompt = f"""你是 Excel 班表解析助手。以下是 Excel 工作表資訊，請分析並回傳解析設定（純 JSON，不加 markdown code block）。

工作表列表：{wb.sheetnames}

各工作表前4行（list of rows，每 row 為 list of cell values）：
{json.dumps(per_sheet, ensure_ascii=False)}

請判斷：
1. schedule_sheet：哪個 sheet 是班表（護理師每日班別），通常每行一位護理師或一筆記錄
2. ot_priority_sheet：哪個 sheet 是加班順位（有順位/order 欄）；無則 null
3. schedule_format："wide"（日期橫向為欄）或 "long"（每行一筆日期+班別）
4. wide_config（format=wide 時）：
   - name_col：姓名欄 index（0-based）
   - area_col：房區/區域欄 index（無則 -1）
   - date_start_col：第一個日期欄 index
   - data_start_row：資料起始行 index（跳過表頭和摘要行，0-based）
   - skip_row_keywords：用來識別非資料行的關鍵字 list（如 ["班人數","合計"]）
5. long_config（format=long 時）：date_col, name_col, shift_col, area_col, data_start_row（均 0-based）
6. shift_map：班別值正規化對應（如 {{"OFF":"休","off":"休","休假":"休","None":""}}）

參考年月：{year}-{month:02d}（日期欄標頭如 "5/1" 請以此年份推算完整日期）

回傳 JSON（不加任何說明文字）：
{{
  "schedule_sheet": "工作表1",
  "ot_priority_sheet": null,
  "schedule_format": "wide",
  "wide_config": {{
    "name_col": 0, "area_col": 3, "date_start_col": 8,
    "data_start_row": 2, "skip_row_keywords": ["班人數"]
  }},
  "long_config": null,
  "shift_map": {{"OFF": "休", "off": "休"}}
}}"""

    try:
        print("[detect_excel_format] calling Claude API...")
        resp = _anthropic_client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=512,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        print(f"[detect_excel_format] response: {raw[:200]}")
        raw = re.sub(r"^```[a-z]*\n?", "", raw).rstrip("`").strip()
        return json.loads(raw)
    except Exception as e:
        print(f"[detect_excel_format] error: {e}")
        return None


OFF_SHIFTS = {"off", "休", "公休", "休假", "例休", "補休", "off日", "休日"}

def _parse_wide_schedule(ws, cfg: dict, shift_map: dict, version: str, year: int, month: int) -> list:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    date_cols = {}
    for ci in range(cfg["date_start_col"], len(headers)):
        raw = headers[ci]
        if not raw:
            continue
        ds = parse_date_header(str(raw), year, month)
        if ds:
            date_cols[ci] = ds

    skip_kws = cfg.get("skip_row_keywords", [])
    name_ci = cfg["name_col"]
    area_ci = cfg.get("area_col", -1)
    data_row = cfg["data_start_row"] + 1  # convert 0-based to openpyxl 1-based

    schedules = []
    for row in ws.iter_rows(min_row=data_row, values_only=True):
        if not any(row):
            continue
        name = str(row[name_ci] or "").strip()
        if not name or any(kw in name for kw in skip_kws):
            continue
        area = str(row[area_ci] or "").strip() if area_ci >= 0 and area_ci < len(row) else ""
        for ci, date_str in date_cols.items():
            if ci >= len(row):
                continue
            raw_shift = str(row[ci] or "").strip()
            shift = shift_map.get(raw_shift, raw_shift)
            if not shift or shift.lower() in OFF_SHIFTS:
                continue
            schedules.append({
                "date_str": date_str,
                "name": name,
                "shift_type": shift,
                "area": area,
                "source_version": version,
            })
    return schedules


def _parse_long_schedule(ws, cfg: dict, shift_map: dict, version: str) -> list:
    date_ci = cfg.get("date_col", 0)
    name_ci = cfg.get("name_col", 1)
    shift_ci = cfg.get("shift_col", 2)
    area_ci = cfg.get("area_col", -1)
    data_row = cfg.get("data_start_row", 1)

    schedules = []
    for row in ws.iter_rows(min_row=data_row + 1, values_only=True):
        if not any(row):
            continue
        raw_date = row[date_ci] if date_ci < len(row) else None
        name = str(row[name_ci] or "").strip() if name_ci < len(row) else ""
        raw_shift = str(row[shift_ci] or "").strip() if shift_ci < len(row) else ""
        shift = shift_map.get(raw_shift, raw_shift)
        area = str(row[area_ci] or "").strip() if area_ci >= 0 and area_ci < len(row) else ""

        if not raw_date or not name or not shift or shift.lower() in OFF_SHIFTS:
            continue
        if isinstance(raw_date, (date, datetime)):
            date_str = raw_date.strftime("%Y-%m-%d") if isinstance(raw_date, datetime) else raw_date.isoformat()
        else:
            date_str = str(raw_date).strip()

        schedules.append({
            "date_str": date_str,
            "name": name,
            "shift_type": shift,
            "area": area,
            "source_version": version,
        })
    return schedules


def _parse_ot_priority_sheet(ws, version: str) -> list:
    """加班順位 sheet 解析（long 格式：日期/順位/姓名/班別）。"""
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}
    ot_priorities = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        raw_date = row[col.get("日期", 0)]
        order = row[col.get("順位", 1)]
        name = str(row[col.get("姓名", 2)] or "").strip()
        shift_type = str(row[col.get("班別", 3)] or "").strip() if len(row) > col.get("班別", 3) else ""
        if not raw_date or order is None or not name or isinstance(order, datetime):
            continue
        if isinstance(raw_date, (date, datetime)):
            date_str = raw_date.strftime("%Y-%m-%d") if isinstance(raw_date, datetime) else raw_date.isoformat()
        else:
            date_str = str(raw_date).strip()
        ot_priorities.append({
            "date_str": date_str,
            "priority_order": int(order),
            "name": name,
            "shift_type": shift_type or None,
            "source_version": version,
        })
    return ot_priorities


def parse_schedule_excel(file_bytes: bytes, version: str = "") -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 優先嘗試已知固定格式（不呼叫 AI）
    known = detect_known_format(wb)
    if known == 'ot_before':
        schedules = _parse_vnhc_wide(wb['本月'], version, strip_ot_code=True) if '本月' in wb.sheetnames else []

        # 7-3：優先用 7_3P sheet
        ot_7_3 = _parse_7_3P_sheet(wb['7_3P'], version) if '7_3P' in wb.sheetnames else []

        # 其他班別：優先用 值班列印全 sheet（排除 7-3）
        ot_others = _parse_duty_print_full_sheet(wb['值班列印全'], version, exclude_shifts={'7-3'}) \
                    if '值班列印全' in wb.sheetnames else []

        # Fallback：本月補齊沒資料的班別
        shifts_covered = {r['shift_type'] for r in ot_7_3 + ot_others}
        ot_fallback = []
        if '本月' in wb.sheetnames:
            all_from_main = _parse_vnhc_ot_priorities(wb['本月'], version)
            ot_fallback = [r for r in all_from_main if r['shift_type'] not in shifts_covered]

        ot_priority = ot_7_3 + ot_others + ot_fallback
        return {
            'schedules': schedules,
            'ot_priority': ot_priority,
            'overtime_records': [],
            '_debug': {
                'format': 'ot_before',
                'parsed_schedule_count': len(schedules),
                'parsed_ot_count': len(ot_priority),
                'ot_7_3_count': len(ot_7_3),
                'ot_others_count': len(ot_others),
                'ot_fallback_count': len(ot_fallback),
            },
        }
    if known == 'ot_after':
        result = _parse_ot_after(wb, version)
        result['ot_priority'] = []
        result['_debug'] = {
            'format': 'ot_after',
            'parsed_schedule_count': len(result['schedules']),
            'parsed_ot_count': 0,
            'parsed_overtime_count': len(result['overtime_records']),
        }
        return result
    if known == 'pure' and '本月' in wb.sheetnames:
        schedules = _parse_vnhc_wide(wb['本月'], version)
        return {
            'schedules': schedules,
            'ot_priority': [],
            'overtime_records': [],
            '_debug': {'format': 'pure', 'parsed_schedule_count': len(schedules), 'parsed_ot_count': 0},
        }

    # Fallback：原有 AI 判讀邏輯
    year, month = infer_year_month(version)
    fmt = detect_excel_format(wb, year, month)
    debug = {
        "format": "ai_fallback",
        "sheets_found": wb.sheetnames,
        "ai_detected": fmt is not None,
        "ai_sheet": fmt.get("schedule_sheet") if fmt else None,
        "ai_format": fmt.get("schedule_format") if fmt else None,
        "ai_ot_sheet": fmt.get("ot_priority_sheet") if fmt else None,
    }

    schedules = []
    ot_priorities = []

    if fmt:
        shift_map = fmt.get("shift_map", {})

        sched_sheet = fmt.get("schedule_sheet")
        if sched_sheet and sched_sheet in wb.sheetnames:
            ws = wb[sched_sheet]
            if fmt.get("schedule_format") == "wide" and fmt.get("wide_config"):
                schedules = _parse_wide_schedule(ws, fmt["wide_config"], shift_map, version, year, month)
            elif fmt.get("schedule_format") == "long" and fmt.get("long_config"):
                schedules = _parse_long_schedule(ws, fmt["long_config"], shift_map, version)
        elif sched_sheet:
            debug["ai_sheet_error"] = f"偵測到的工作表「{sched_sheet}」不存在於檔案中"
    else:
        debug["ai_error"] = "AI 判讀失敗，使用舊版固定格式解析"
        if "班表" in wb.sheetnames:
            ws = wb["班表"]
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col = {h: i for i, h in enumerate(headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                raw_date = row[col.get("日期", 0)]
                name = str(row[col.get("姓名", 1)] or "").strip()
                shift = str(row[col.get("班別", 2)] or "").strip()
                area = str(row[col.get("房區", 3)] or "").strip() if "房區" in col else ""
                if not raw_date or not name or not shift:
                    continue
                if isinstance(raw_date, (date, datetime)):
                    date_str = raw_date.strftime("%Y-%m-%d") if isinstance(raw_date, datetime) else raw_date.isoformat()
                else:
                    date_str = str(raw_date).strip()
                schedules.append({"date_str": date_str, "name": name, "shift_type": shift, "area": area, "source_version": version})

        if "加班順位" in wb.sheetnames:
            ot_priorities = _parse_ot_priority_sheet(wb["加班順位"], version)

    debug["parsed_schedule_count"] = len(schedules)
    debug["parsed_ot_count"] = len(ot_priorities)
    return {"schedules": schedules, "ot_priority": ot_priorities, "overtime_records": [], "_debug": debug}


# ── Date utilities ─────────────────────────────────────────────────────────────

def week_range_of(d: date):
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    return monday.isoformat(), sunday.isoformat()


def month_range(year_month: str):
    y, m = int(year_month[:4]), int(year_month[5:7])
    first = date(y, m, 1)
    if m == 12:
        last = date(y + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(y, m + 1, 1) - timedelta(days=1)
    return first.isoformat(), last.isoformat()


# ── Formatting ─────────────────────────────────────────────────────────────────

def format_schedule_day(rows: list, all_users: dict) -> str:
    if not rows:
        return "（查無班表資料）"
    lines = []
    for r in rows:
        name = all_users.get(r["user_id"], {}).get("name", "?")
        shift = r["shift_type"]
        area = f" {r['area']}" if r.get("area") else ""
        lines.append(f"  {name}：{shift}{area}")
    return "\n".join(lines)


def format_own_schedule(rows: list, all_users: dict) -> str:
    if not rows:
        return "（查無班表資料）"
    lines = []
    for r in rows:
        name = all_users.get(r["user_id"], {}).get("name", "?")
        shift = r["shift_type"]
        area = f" {r['area']}" if r.get("area") else ""
        lines.append(f"  {r['schedule_date']} {name}：{shift}{area}")
    return "\n".join(lines)


def format_ot_priority(rows: list, all_users: dict) -> str:
    if not rows:
        return "（查無加班順位資料）"
    lines = []
    current_shift = None
    for r in rows:
        name = all_users.get(r["user_id"], {}).get("name", "?")
        shift = r.get("shift_type") or "—"
        if shift != current_shift:
            lines.append(f"【{shift}】")
            current_shift = shift
        lines.append(f"  {r['priority_order']}. {name}")
    return "\n".join(lines)


def format_swap_request_line(req: dict, all_users: dict) -> str:
    rtype = "換班" if req["request_type"] == "shift" else ("直接調班" if req["request_type"] == "manager_direct" else "換加班順位")
    status_map = {
        "submitted": "已提交",
        "pending_peer": "等待對方確認",
        "pending_admin": "等待主管審核",
        "approved": "已核准",
        "rejected": "已拒絕",
        "cancelled": "已取消",
        "conflict_rejected": "因衝突自動取消",
        "peer_rejected": "對方已拒絕",
    }
    status = status_map.get(req["status"], req["status"])
    req_name = all_users.get(req.get("requester_id"), {}).get("name", "?") if req.get("requester_id") else "系統"
    tgt_name = all_users.get(req.get("target_user_id"), {}).get("name", "?") if req.get("target_user_id") else "未指定"
    short_id = str(req["id"])[:8]
    return f"[{short_id}] {rtype} {req_name}→{tgt_name} ({status})"


def _build_users_map():
    users = get_all_users()
    return {u["id"]: u for u in users}


# ── Auth dependencies ──────────────────────────────────────────────────────────

async def get_current_user(request: Request):
    line_user_id = request.headers.get("X-Line-User-Id")
    if not line_user_id:
        raise HTTPException(status_code=401, detail="Missing X-Line-User-Id header")
    user = get_user_by_line_id(line_user_id)
    if not user:
        raise HTTPException(status_code=403, detail="User not bound")
    if user["role"] == "pending":
        raise HTTPException(status_code=403, detail="Account pending approval")
    return user


def require_manager(user=Depends(get_current_user)):
    if user["role"] not in ("manager", "admin"):
        raise HTTPException(status_code=403, detail="Manager or admin required")
    return user


# ── Pydantic models ────────────────────────────────────────────────────────────

class SwapRequestCreate(BaseModel):
    request_type: str  # shift | ot_priority
    schedule_id: Optional[str] = None
    ot_priority_id: Optional[str] = None
    target_user_id: str
    target_schedule_id: Optional[str] = None
    target_ot_priority_id: Optional[str] = None
    reason: Optional[str] = None


class SwapRespond(BaseModel):
    response: str  # accepted | rejected


class SwapReview(BaseModel):
    decision: str  # approved | rejected
    comment: Optional[str] = None


class DirectSwapBody(BaseModel):
    swap_type: str  # shift | ot_priority
    slot_a_id: str
    slot_b_id: str
    reason: Optional[str] = None


# ── REST API endpoints ─────────────────────────────────────────────────────────

@app.get("/")
async def root():
    return {"service": "nurse-schedule-line", "status": "ok"}


@app.get("/health")
async def health():
    issues = []
    if not LINE_TOKEN:
        issues.append("LINE_CHANNEL_ACCESS_TOKEN missing")
    if not SUPABASE_URL or not SUPABASE_KEY:
        issues.append("Supabase credentials missing")
    if issues:
        return JSONResponse(status_code=500, content={"status": "error", "issues": issues})
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/nurses",
            headers=SUPABASE_HEADERS,
            params={"limit": "1"},
            timeout=5,
        )
        db_ok = r.status_code < 400
    except Exception:
        db_ok = False
    if not db_ok:
        return JSONResponse(status_code=500, content={"status": "error", "issues": ["DB unreachable"]})
    return {"status": "ok"}


@app.get("/api/me")
async def api_me(user=Depends(get_current_user)):
    return {"id": user["id"], "name": user["name"], "role": user["role"]}


@app.get("/api/schedules/me")
async def api_schedules_me(
    mode: str = "week",
    start: Optional[str] = None,
    end: Optional[str] = None,
    user=Depends(get_current_user),
):
    today = date.today()
    if mode == "today":
        s, e = today.isoformat(), today.isoformat()
    elif mode == "tomorrow":
        t = today + timedelta(days=1)
        s, e = t.isoformat(), t.isoformat()
    elif mode == "week":
        s, e = week_range_of(today)
    elif mode == "month":
        s, e = month_range(today.strftime("%Y-%m"))
    elif mode == "range" and start and end:
        s, e = start, end
    else:
        s, e = week_range_of(today)

    rows = _get_schedules_range(s, e, user_id=user["id"])
    users_map = _build_users_map()
    return {"schedules": rows, "users": {uid: u["name"] for uid, u in users_map.items()}}


@app.get("/api/schedules")
async def api_schedules(
    date: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    name: Optional[str] = None,
    user=Depends(get_current_user),
):
    today_str = datetime.today().strftime("%Y-%m-%d")
    if date:
        rows = get_schedules_by_date(date)
    elif start and end:
        rows = _get_schedules_range(start, end)
    else:
        rows = get_schedules_by_date(today_str)
    users_map = _build_users_map()
    if name:
        kw = name.strip().lower()
        uid_set = {uid for uid, u in users_map.items() if kw in u["name"].lower()}
        rows = [r for r in rows if r["user_id"] in uid_set]
    return {"schedules": rows, "users": {uid: u["name"] for uid, u in users_map.items()}}


@app.get("/api/ot-priority/me")
async def api_ot_priority_me(
    month: Optional[str] = None,
    user=Depends(get_current_user),
):
    if month:
        start, end = month_range(month)
    else:
        start, end = month_range(datetime.today().strftime("%Y-%m"))
    rows = get_ot_priority_range(start, end, user_id=user["id"])
    return {"ot_priority": rows}


@app.get("/api/ot-priority")
async def api_ot_priority(
    date: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    shift_type: Optional[str] = None,
    name: Optional[str] = None,
    user=Depends(get_current_user),
):
    users_map = _build_users_map()
    if start and end:
        rows = get_ot_priority_range(start, end)
    else:
        target_date = date or datetime.today().strftime("%Y-%m-%d")
        rows = get_ot_priority_by_date(target_date, shift_type)
    if name:
        kw = name.strip().lower()
        uid_set = {uid for uid, u in users_map.items() if kw in u["name"].lower()}
        rows = [r for r in rows if r["user_id"] in uid_set]
    return {"ot_priority": rows, "users": {uid: u["name"] for uid, u in users_map.items()}}


@app.post("/api/schedules/import")
async def api_import_schedules(
    file: UploadFile = File(...),
    version: str = Form(""),
    import_schedules: bool = Form(True),
    import_ot_priority: bool = Form(True),
    start_date: str = Form(""),
    end_date: str = Form(""),
    manager=Depends(require_manager),
):
    content = await file.read()
    if not version:
        version = file.filename.replace(".xlsm", "").replace(".xlsx", "").replace(".xls", "")

    parsed = parse_schedule_excel(content, version)
    debug_info = parsed.get("_debug", {})
    users_map = {u["name"]: u["id"] for u in get_all_users()}

    schedule_records = []
    ot_records = []
    unmatched_names = set()
    invalid_shifts = set()
    all_dates = []

    for item in parsed["schedules"]:
        all_dates.append(item["date_str"])
        if start_date and item["date_str"] < start_date:
            continue
        if end_date and item["date_str"] > end_date:
            continue
        uid = users_map.get(item["name"])
        if not uid:
            unmatched_names.add(item["name"])
            continue
        if not item["shift_type"] or item["shift_type"].lower() in OFF_SHIFTS:
            invalid_shifts.add(item["shift_type"])
            continue
        schedule_records.append({
            "user_id": uid,
            "schedule_date": item["date_str"],
            "shift_type": item["shift_type"],
            "area": item["area"] or None,
            "source_version": item["source_version"],
            "status": "active",
        })

    for item in parsed["ot_priority"]:
        all_dates.append(item["date_str"])
        if start_date and item["date_str"] < start_date:
            continue
        if end_date and item["date_str"] > end_date:
            continue
        uid = users_map.get(item["name"])
        if not uid:
            unmatched_names.add(item["name"])
            continue
        ot_records.append({
            "priority_date": item["date_str"],
            "user_id": uid,
            "shift_type": item.get("shift_type"),
            "priority_order": item["priority_order"],
            "source_version": item["source_version"],
            "status": "active",
        })

    sched_ok = True
    ot_ok = True
    sched_inserted = 0
    ot_inserted = 0

    if schedule_records and import_schedules:
        delete_schedules_by_version(version)
        sched_ok = upsert_schedules(schedule_records)
        sched_inserted = len(schedule_records) if sched_ok else 0

    if ot_records and import_ot_priority:
        delete_ot_priority_by_version(version)
        ot_ok = upsert_ot_priority(ot_records)
        ot_inserted = len(ot_records) if ot_ok else 0

    return {
        "schedules_imported": sched_inserted,
        "ot_priority_imported": ot_inserted,
        "parsed_schedules": len(schedule_records),
        "parsed_ot_priority": len(ot_records),
        "parsed_start": min(all_dates) if all_dates else "",
        "parsed_end": max(all_dates) if all_dates else "",
        "dry_run": not import_schedules and not import_ot_priority,
        "unmatched_names": list(unmatched_names),
        "invalid_shifts": list(invalid_shifts),
        "success": sched_ok and ot_ok,
        "debug": debug_info,
    }


@app.post("/api/swap-requests")
async def api_create_swap_request(
    body: SwapRequestCreate,
    user=Depends(get_current_user),
):
    if body.request_type not in ("shift", "ot_priority"):
        raise HTTPException(status_code=400, detail="request_type must be shift or ot_priority")

    if body.request_type == "shift":
        if not body.schedule_id or not body.target_schedule_id:
            raise HTTPException(status_code=400, detail="schedule_id and target_schedule_id required")
        my_slot = get_schedule_by_id(body.schedule_id)
        their_slot = get_schedule_by_id(body.target_schedule_id)
        if not my_slot or my_slot["user_id"] != user["id"]:
            raise HTTPException(status_code=400, detail="schedule_id does not belong to you")
        if not their_slot or their_slot["user_id"] != body.target_user_id:
            raise HTTPException(status_code=400, detail="target_schedule_id does not belong to target_user_id")
    else:
        if not body.ot_priority_id or not body.target_ot_priority_id:
            raise HTTPException(status_code=400, detail="ot_priority_id and target_ot_priority_id required")
        my_slot = get_ot_priority_by_id(body.ot_priority_id)
        their_slot = get_ot_priority_by_id(body.target_ot_priority_id)
        if not my_slot or my_slot["user_id"] != user["id"]:
            raise HTTPException(status_code=400, detail="ot_priority_id does not belong to you")
        if not their_slot or their_slot["user_id"] != body.target_user_id:
            raise HTTPException(status_code=400, detail="target_ot_priority_id does not belong to target_user_id")
        if my_slot.get("shift_type") != their_slot.get("shift_type"):
            raise HTTPException(status_code=400, detail=f"班別不同（{my_slot.get('shift_type')} vs {their_slot.get('shift_type')}），不可交換加班順位")

    data = {
        "request_type": body.request_type,
        "status": "pending_peer",
        "requester_id": user["id"],
        "target_user_id": body.target_user_id,
        "reason": body.reason,
    }
    if body.request_type == "shift":
        data["schedule_id"] = body.schedule_id
        data["target_schedule_id"] = body.target_schedule_id
    else:
        data["ot_priority_id"] = body.ot_priority_id
        data["target_ot_priority_id"] = body.target_ot_priority_id

    req = create_swap_request(data)
    if not req:
        raise HTTPException(status_code=500, detail="Failed to create swap request")

    target = get_user_by_id(body.target_user_id)
    if target and target.get("line_user_id"):
        rtype = "換班" if body.request_type == "shift" else "換加班順位"
        push_message(
            target["line_user_id"],
            f"【{rtype}請求】{user['name']} 希望與您換班，請至 LINE app 查看並確認。\n申請 ID：{str(req['id'])[:8]}",
        )

    return req


@app.get("/api/swap-requests/me")
async def api_my_swap_requests(user=Depends(get_current_user)):
    reqs = get_swap_requests_by_user(user["id"])
    users_map = _build_users_map()
    return {
        "swap_requests": reqs,
        "users": {uid: u["name"] for uid, u in users_map.items()},
    }


@app.post("/api/swap-requests/{request_id}/cancel")
async def api_cancel_swap_request(request_id: str, user=Depends(get_current_user)):
    req = get_swap_request_by_id(request_id)
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["requester_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Not your request")
    if req["status"] not in ("submitted", "pending_peer"):
        raise HTTPException(status_code=400, detail=f"Cannot cancel request in status {req['status']}")
    update_swap_request(request_id, {"status": "cancelled"})
    return {"status": "cancelled"}


@app.post("/api/swap-requests/{request_id}/respond")
async def api_respond_swap_request(
    request_id: str,
    body: SwapRespond,
    user=Depends(get_current_user),
):
    req = get_swap_request_by_id(request_id)
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["target_user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Not your request to respond")
    if req["status"] != "pending_peer":
        raise HTTPException(status_code=400, detail=f"Request is not awaiting peer response (status: {req['status']})")

    if body.response == "rejected":
        update_swap_request(request_id, {
            "status": "peer_rejected",
            "peer_response": "rejected",
            "peer_responded_at": datetime.utcnow().isoformat(),
        })
        requester = get_user_by_id(req["requester_id"])
        if requester and requester.get("line_user_id"):
            push_message(requester["line_user_id"], f"您的換班申請 [{str(request_id)[:8]}] 已被對方拒絕。")
        return {"status": "peer_rejected"}

    if body.response != "accepted":
        raise HTTPException(status_code=400, detail="response must be accepted or rejected")

    if req["request_type"] == "ot_priority":
        # OT priority swap: execute immediately
        ot_a = get_ot_priority_by_id(req["ot_priority_id"])   # requester's slot
        ot_b = get_ot_priority_by_id(req["target_ot_priority_id"])  # target's slot
        if not ot_a or not ot_b:
            raise HTTPException(status_code=400, detail="OT priority slots not found")

        # Detect conflicts BEFORE swapping (cross-date: A may already have slot on B's date)
        conflicts_a = _sb("ot_priority", params={
            "user_id": f"eq.{req['requester_id']}",
            "priority_date": f"eq.{ot_b['priority_date']}",
            "shift_type": f"eq.{ot_b['shift_type']}",
            "status": "eq.active",
        }) or []
        conflict_a = next((r for r in conflicts_a if r["id"] != ot_b["id"]), None)

        conflicts_b = _sb("ot_priority", params={
            "user_id": f"eq.{req['target_user_id']}",
            "priority_date": f"eq.{ot_a['priority_date']}",
            "shift_type": f"eq.{ot_a['shift_type']}",
            "status": "eq.active",
        }) or []
        conflict_b = next((r for r in conflicts_b if r["id"] != ot_a["id"]), None)

        # Main swap
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/ot_priority",
            headers=SUPABASE_HEADERS,
            params={"id": f"eq.{ot_a['id']}"},
            json={"user_id": ot_b["user_id"]},
            timeout=10,
        )
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/ot_priority",
            headers=SUPABASE_HEADERS,
            params={"id": f"eq.{ot_b['id']}"},
            json={"user_id": ot_a["user_id"]},
            timeout=10,
        )

        # Secondary swap to resolve duplicate slots caused by cross-date swap
        secondary_msg = ""
        if conflict_a and conflict_b:
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/ot_priority",
                headers=SUPABASE_HEADERS,
                params={"id": f"eq.{conflict_a['id']}"},
                json={"user_id": req["target_user_id"]},
                timeout=10,
            )
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/ot_priority",
                headers=SUPABASE_HEADERS,
                params={"id": f"eq.{conflict_b['id']}"},
                json={"user_id": req["requester_id"]},
                timeout=10,
            )
            secondary_msg = f"（同時已自動對調 {ot_b['priority_date']} 第{conflict_a['priority_order']}順位 與 {ot_a['priority_date']} 第{conflict_b['priority_order']}順位）"

        update_swap_request(request_id, {
            "status": "approved",
            "peer_response": "accepted",
            "peer_responded_at": datetime.utcnow().isoformat(),
        })
        auto_reject_conflicts(
            request_id,
            schedule_ids=[],
            ot_priority_ids=[ot_a["id"], ot_b["id"]],
        )

        # Notify requester
        requester = get_user_by_id(req["requester_id"])
        if requester and requester.get("line_user_id"):
            push_message(requester["line_user_id"], f"您的換加班順位申請 [{str(request_id)[:8]}] 已完成，對方已同意並執行。{secondary_msg}")
        push_message(user["line_user_id"], f"換加班順位申請 [{str(request_id)[:8]}] 已完成。{secondary_msg}")
        return {"status": "approved"}

    else:
        # Shift swap: peer accepted → send to manager
        update_swap_request(request_id, {
            "status": "pending_admin",
            "peer_response": "accepted",
            "peer_responded_at": datetime.utcnow().isoformat(),
        })
        requester = get_user_by_id(req["requester_id"])
        if requester and requester.get("line_user_id"):
            push_message(requester["line_user_id"], f"您的換班申請 [{str(request_id)[:8]}] 對方已同意，等待主管審核。")
        for reviewer in get_reviewers():
            if reviewer.get("line_user_id"):
                push_message(reviewer["line_user_id"], f"【換班審核】有新的換班申請待審核，申請 ID：{str(request_id)[:8]}。")
        return {"status": "pending_admin"}


@app.get("/api/swap-requests/pending")
async def api_pending_swap_requests(manager=Depends(require_manager)):
    reqs = get_pending_swap_requests()
    users_map = _build_users_map()
    return {
        "swap_requests": reqs,
        "users": {uid: u["name"] for uid, u in users_map.items()},
    }


@app.post("/api/swap-requests/{request_id}/review")
async def api_review_swap_request(
    request_id: str,
    body: SwapReview,
    manager=Depends(require_manager),
):
    req = get_swap_request_by_id(request_id)
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["status"] != "pending_admin":
        raise HTTPException(status_code=400, detail=f"Request is not pending admin review (status: {req['status']})")
    if req["request_type"] != "shift":
        raise HTTPException(status_code=400, detail="Only shift swaps require admin review")

    if body.decision == "rejected":
        update_swap_request(request_id, {
            "status": "rejected",
            "admin_decision": "rejected",
            "admin_comment": body.comment,
            "admin_decided_at": datetime.utcnow().isoformat(),
        })
        for uid in [req.get("requester_id"), req.get("target_user_id")]:
            if uid:
                person = get_user_by_id(uid)
                if person and person.get("line_user_id"):
                    push_message(person["line_user_id"], f"換班申請 [{str(request_id)[:8]}] 已被主管拒絕。{body.comment or ''}")
        return {"status": "rejected"}

    if body.decision != "approved":
        raise HTTPException(status_code=400, detail="decision must be approved or rejected")

    sched_a = get_schedule_by_id(req["schedule_id"])
    sched_b = get_schedule_by_id(req["target_schedule_id"])
    if not sched_a or not sched_b:
        raise HTTPException(status_code=400, detail="Schedule slots not found")

    requests.patch(
        f"{SUPABASE_URL}/rest/v1/schedules",
        headers=SUPABASE_HEADERS,
        params={"id": f"eq.{sched_a['id']}"},
        json={"user_id": sched_b["user_id"]},
        timeout=10,
    )
    requests.patch(
        f"{SUPABASE_URL}/rest/v1/schedules",
        headers=SUPABASE_HEADERS,
        params={"id": f"eq.{sched_b['id']}"},
        json={"user_id": sched_a["user_id"]},
        timeout=10,
    )
    update_swap_request(request_id, {
        "status": "approved",
        "admin_decision": "approved",
        "admin_comment": body.comment,
        "admin_decided_at": datetime.utcnow().isoformat(),
    })
    auto_reject_conflicts(
        request_id,
        schedule_ids=[sched_a["id"], sched_b["id"]],
        ot_priority_ids=[],
    )

    for uid in [req.get("requester_id"), req.get("target_user_id")]:
        if uid:
            person = get_user_by_id(uid)
            if person and person.get("line_user_id"):
                push_message(person["line_user_id"], f"換班申請 [{str(request_id)[:8]}] 已獲主管核准，班表已更新。")

    return {"status": "approved"}


@app.post("/api/admin/direct-swap")
async def api_direct_swap(body: DirectSwapBody, manager=Depends(require_manager)):
    if body.swap_type == "shift":
        slot_a = get_schedule_by_id(body.slot_a_id)
        slot_b = get_schedule_by_id(body.slot_b_id)
        if not slot_a or not slot_b:
            raise HTTPException(status_code=404, detail="Schedule slot(s) not found")

        requests.patch(
            f"{SUPABASE_URL}/rest/v1/schedules",
            headers=SUPABASE_HEADERS,
            params={"id": f"eq.{slot_a['id']}"},
            json={"user_id": slot_b["user_id"]},
            timeout=10,
        )
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/schedules",
            headers=SUPABASE_HEADERS,
            params={"id": f"eq.{slot_b['id']}"},
            json={"user_id": slot_a["user_id"]},
            timeout=10,
        )
        create_swap_request({
            "request_type": "manager_direct",
            "status": "approved",
            "schedule_id": slot_a["id"],
            "target_schedule_id": slot_b["id"],
            "target_user_id": slot_b["user_id"],
            "reason": body.reason,
            "performed_by": manager["id"],
            "admin_decision": "approved",
            "admin_decided_at": datetime.utcnow().isoformat(),
        })
        auto_reject_conflicts("", schedule_ids=[slot_a["id"], slot_b["id"]], ot_priority_ids=[])
        for uid in [slot_a["user_id"], slot_b["user_id"]]:
            person = get_user_by_id(uid)
            if person and person.get("line_user_id"):
                push_message(person["line_user_id"], f"主管已直接調整您的班表，請至 LINE app 查看最新班表。")
        return {"status": "done"}

    elif body.swap_type == "ot_priority":
        slot_a = get_ot_priority_by_id(body.slot_a_id)
        slot_b = get_ot_priority_by_id(body.slot_b_id)
        if not slot_a or not slot_b:
            raise HTTPException(status_code=404, detail="OT priority slot(s) not found")

        requests.patch(
            f"{SUPABASE_URL}/rest/v1/ot_priority",
            headers=SUPABASE_HEADERS,
            params={"id": f"eq.{slot_a['id']}"},
            json={"user_id": slot_b["user_id"]},
            timeout=10,
        )
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/ot_priority",
            headers=SUPABASE_HEADERS,
            params={"id": f"eq.{slot_b['id']}"},
            json={"user_id": slot_a["user_id"]},
            timeout=10,
        )
        create_swap_request({
            "request_type": "manager_direct",
            "status": "approved",
            "ot_priority_id": slot_a["id"],
            "target_ot_priority_id": slot_b["id"],
            "target_user_id": slot_b["user_id"],
            "reason": body.reason,
            "performed_by": manager["id"],
            "admin_decision": "approved",
            "admin_decided_at": datetime.utcnow().isoformat(),
        })
        auto_reject_conflicts("", schedule_ids=[], ot_priority_ids=[slot_a["id"], slot_b["id"]])
        for uid in [slot_a["user_id"], slot_b["user_id"]]:
            person = get_user_by_id(uid)
            if person and person.get("line_user_id"):
                push_message(person["line_user_id"], f"主管已直接調整您的加班順位，請至 LINE app 查看最新資料。")
        return {"status": "done"}

    raise HTTPException(status_code=400, detail="swap_type must be shift or ot_priority")


# ── Static file serving ────────────────────────────────────────────────────────

@app.get("/{filename}.html")
async def serve_html(filename: str):
    path = os.path.join(DOCS_DIR, f"{filename}.html")
    if not os.path.exists(path):
        raise HTTPException(status_code=404)
    content = open(path, encoding="utf-8").read()
    content = content.replace("{{LIFF_ID}}", LIFF_ID)
    content = content.replace("{{API_BASE}}", API_BASE)
    return Response(content, media_type="text/html")


@app.get("/share.js")
async def serve_js():
    path = os.path.join(DOCS_DIR, "share.js")
    if os.path.exists(path):
        return FileResponse(path, media_type="application/javascript")
    raise HTTPException(status_code=404)


@app.get("/{filename}.png")
async def serve_png(filename: str):
    path = os.path.join(DOCS_DIR, f"{filename}.png")
    if os.path.exists(path):
        return FileResponse(path, media_type="image/png")
    raise HTTPException(status_code=404)


# ── LINE Webhook ───────────────────────────────────────────────────────────────

def verify_line_signature(body_bytes: bytes, signature: str) -> bool:
    if not LINE_SECRET:
        return True
    digest = hmac.new(LINE_SECRET.encode(), body_bytes, hashlib.sha256).digest()
    expected = base64.b64encode(digest).decode()
    return hmac.compare_digest(expected, signature)


@app.post("/webhook")
async def webhook(request: Request):
    body_bytes = await request.body()
    sig = request.headers.get("X-Line-Signature", "")
    if not verify_line_signature(body_bytes, sig):
        raise HTTPException(status_code=400, detail="Invalid signature")

    body = json.loads(body_bytes)
    for event in body.get("events", []):
        _handle_event(event)
    return {"status": "ok"}


def _handle_event(event: dict):
    etype = event.get("type")
    reply_token = event.get("replyToken", "")

    if etype == "follow":
        reply_message(reply_token, (
            "歡迎使用護理排班系統！\n\n"
            "常用指令：\n"
            "・我的班表 → 本週班表\n"
            "・今天班表 → 今日全員班表\n"
            "・明天班表 → 明日全員班表\n"
            "・查班表 MM/DD → 指定日期全員班表\n"
            "・加班順位 今天 → 今日加班順位\n"
            "・加班順位 MM/DD → 指定日期加班順位\n"
            "・我的申請 → 查看換班申請\n"
            "・取消申請 [ID] → 取消申請\n\n"
            "換班功能請使用 LINE app 操作。"
        ))
        return

    if etype != "message":
        return

    msg = event.get("message", {})
    if msg.get("type") != "text":
        return

    text = msg.get("text", "").strip()
    source = event.get("source", {})
    line_user_id = source.get("userId", "")

    if not line_user_id:
        return

    user = get_user_by_line_id(line_user_id)
    if not user or user["role"] == "pending":
        reply_message(reply_token, "您的帳號尚未綁定或待審核，請聯絡管理者。")
        return

    _dispatch_command(text, user, reply_token)


def _dispatch_command(text: str, user: dict, reply_token: str):
    today = date.today()
    users_map = _build_users_map()

    # ── 我的班表
    if text == "我的班表":
        s, e = week_range_of(today)
        rows = _get_schedules_range(s, e, user_id=user["id"])
        body = format_own_schedule(rows, users_map)
        reply_message(reply_token, f"本週班表（{s} ~ {e}）：\n{body}")
        return

    # ── 本週班表
    if text == "本週班表":
        s, e = week_range_of(today)
        rows = _get_schedules_range(s, e, user_id=user["id"])
        body = format_own_schedule(rows, users_map)
        reply_message(reply_token, f"本週班表（{s} ~ {e}）：\n{body}")
        return

    # ── 今天班表
    if text == "今天班表":
        rows = get_schedules_by_date(today.isoformat())
        body = format_schedule_day(rows, users_map)
        reply_message(reply_token, f"今天（{today.isoformat()}）全員班表：\n{body}")
        return

    # ── 明天班表
    if text == "明天班表":
        tomorrow = (today + timedelta(days=1)).isoformat()
        rows = get_schedules_by_date(tomorrow)
        body = format_schedule_day(rows, users_map)
        reply_message(reply_token, f"明天（{tomorrow}）全員班表：\n{body}")
        return

    # ── 查班表 MM/DD 或 YYYY-MM-DD
    if text.startswith("查班表"):
        parts = text.split()
        if len(parts) < 2:
            reply_message(reply_token, "格式：查班表 MM/DD 或 查班表 YYYY-MM-DD")
            return
        date_str = _parse_date_arg(parts[1], today)
        if not date_str:
            reply_message(reply_token, "日期格式錯誤，請用 MM/DD 或 YYYY-MM-DD。")
            return
        rows = get_schedules_by_date(date_str)
        body = format_schedule_day(rows, users_map)
        reply_message(reply_token, f"{date_str} 全員班表：\n{body}")
        return

    # ── 加班順位 今天 / MM/DD / YYYY-MM-DD
    if text.startswith("加班順位"):
        parts = text.split()
        if len(parts) < 2 or parts[1] == "今天":
            target_date = today.isoformat()
        else:
            target_date = _parse_date_arg(parts[1], today)
            if not target_date:
                reply_message(reply_token, "日期格式錯誤，請用 今天、MM/DD 或 YYYY-MM-DD。")
                return
        rows = get_ot_priority_by_date(target_date)
        body = format_ot_priority(rows, users_map)
        reply_message(reply_token, f"{target_date} 加班順位：\n{body}")
        return

    # ── 我的申請
    if text == "我的申請":
        reqs = get_swap_requests_by_user(user["id"])
        if not reqs:
            reply_message(reply_token, "目前無換班申請記錄。")
            return
        lines = [format_swap_request_line(r, users_map) for r in reqs[:10]]
        reply_message(reply_token, "近期換班申請：\n" + "\n".join(lines))
        return

    # ── 取消申請 [ID]
    if text.startswith("取消申請"):
        parts = text.split()
        if len(parts) < 2:
            reply_message(reply_token, "格式：取消申請 [申請ID前8碼]")
            return
        short_id = parts[1].strip()
        matched = _find_request_by_short_id(short_id, user["id"])
        if not matched:
            reply_message(reply_token, f"找不到申請 {short_id}，或不屬於您。")
            return
        if matched["status"] not in ("submitted", "pending_peer"):
            reply_message(reply_token, f"申請目前狀態為「{matched['status']}」，無法取消。")
            return
        update_swap_request(matched["id"], {"status": "cancelled"})
        reply_message(reply_token, f"申請 {short_id} 已取消。")
        return

    # ── Manager: 待審換班
    if text == "待審換班" and user["role"] in ("manager", "admin"):
        reqs = get_pending_swap_requests()
        if not reqs:
            reply_message(reply_token, "目前無待審換班申請。")
            return
        lines = [format_swap_request_line(r, users_map) for r in reqs[:10]]
        reply_message(reply_token, "待審換班申請：\n" + "\n".join(lines) + "\n\n回覆「核准換班 [ID]」或「拒絕換班 [ID] [原因]」")
        return

    # ── Manager: 核准換班 [ID]
    if text.startswith("核准換班") and user["role"] in ("manager", "admin"):
        parts = text.split(maxsplit=1)
        if len(parts) < 2:
            reply_message(reply_token, "格式：核准換班 [申請ID前8碼]")
            return
        short_id = parts[1].strip()
        req = _find_swap_request_by_short_id_any(short_id)
        if not req or req["status"] != "pending_admin":
            reply_message(reply_token, f"找不到待審核申請 {short_id}。")
            return
        _execute_shift_swap_approval(req, user, comment=None)
        reply_message(reply_token, f"換班申請 {short_id} 已核准，班表已更新。")
        return

    # ── Manager: 拒絕換班 [ID] [原因]
    if text.startswith("拒絕換班") and user["role"] in ("manager", "admin"):
        parts = text.split(maxsplit=2)
        if len(parts) < 2:
            reply_message(reply_token, "格式：拒絕換班 [申請ID前8碼] [原因(選填)]")
            return
        short_id = parts[1].strip()
        comment = parts[2].strip() if len(parts) > 2 else ""
        req = _find_swap_request_by_short_id_any(short_id)
        if not req or req["status"] != "pending_admin":
            reply_message(reply_token, f"找不到待審核申請 {short_id}。")
            return
        update_swap_request(req["id"], {
            "status": "rejected",
            "admin_decision": "rejected",
            "admin_comment": comment,
            "admin_decided_at": datetime.utcnow().isoformat(),
        })
        for uid in [req.get("requester_id"), req.get("target_user_id")]:
            if uid:
                person = get_user_by_id(uid)
                if person and person.get("line_user_id"):
                    push_message(person["line_user_id"], f"換班申請 [{short_id}] 已被主管拒絕。{comment}")
        reply_message(reply_token, f"換班申請 {short_id} 已拒絕。")
        return

    reply_message(reply_token, (
        "指令不認識，常用指令：\n"
        "・我的班表\n"
        "・今天班表 / 明天班表\n"
        "・查班表 MM/DD\n"
        "・加班順位 今天\n"
        "・我的申請\n"
        "・取消申請 [ID]"
    ))


def _parse_date_arg(s: str, today: date) -> Optional[str]:
    s = s.strip()
    if "/" in s and len(s) <= 5:
        parts = s.split("/")
        try:
            return date(today.year, int(parts[0]), int(parts[1])).isoformat()
        except Exception:
            return None
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return s
    return None


def _find_request_by_short_id(short_id: str, user_id: str):
    reqs = get_swap_requests_by_user(user_id)
    for r in reqs:
        if str(r["id"]).startswith(short_id):
            return r
    return None


def _find_swap_request_by_short_id_any(short_id: str):
    rows = _sb("swap_requests", params={
        "status": "eq.pending_admin",
        "order": "created_at.desc",
        "limit": "100",
    }) or []
    for r in rows:
        if str(r["id"]).startswith(short_id):
            return r
    return None


def _execute_shift_swap_approval(req: dict, manager: dict, comment: Optional[str]):
    sched_a = get_schedule_by_id(req["schedule_id"])
    sched_b = get_schedule_by_id(req["target_schedule_id"])
    if not sched_a or not sched_b:
        return

    requests.patch(
        f"{SUPABASE_URL}/rest/v1/schedules",
        headers=SUPABASE_HEADERS,
        params={"id": f"eq.{sched_a['id']}"},
        json={"user_id": sched_b["user_id"]},
        timeout=10,
    )
    requests.patch(
        f"{SUPABASE_URL}/rest/v1/schedules",
        headers=SUPABASE_HEADERS,
        params={"id": f"eq.{sched_b['id']}"},
        json={"user_id": sched_a["user_id"]},
        timeout=10,
    )
    update_swap_request(req["id"], {
        "status": "approved",
        "admin_decision": "approved",
        "admin_comment": comment,
        "admin_decided_at": datetime.utcnow().isoformat(),
    })
    auto_reject_conflicts(req["id"], schedule_ids=[sched_a["id"], sched_b["id"]], ot_priority_ids=[])
    for uid in [req.get("requester_id"), req.get("target_user_id")]:
        if uid:
            person = get_user_by_id(uid)
            if person and person.get("line_user_id"):
                push_message(person["line_user_id"], f"換班申請 [{str(req['id'])[:8]}] 已獲主管核准，班表已更新。")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=PORT, reload=True)
